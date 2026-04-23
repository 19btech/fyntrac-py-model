"""
diagnose_tnt001.py
==================
End-to-end diagnostic that runs the full pipeline locally:
  1. Connect to MongoDB TNT001
  2. Load active Model -> ModelFile (BSON Binary)
  3. Decode Python/Excel model file
  4. Fetch EventHistory events
  5. Run data_transformer
  6. Execute model via ModelRunner
  7. Print generated transactions

Run from the project root:
    source venv/bin/activate
    python diagnose_tnt001.py
"""

import io
import json
import sys
import pymongo
from bson import ObjectId

MONGO_URI = "mongodb://root:R3s3rv%23313@127.0.0.1:27017/?authSource=admin"
TENANT    = "TNT001"    # change to TNT002 etc. as needed
POSTING_DATE = None     # set to an int like 20220131 to filter, or None for all

# ─────────────────────────────────────────────────────────────────────────────
print(f"\n{'='*60}")
print(f"  Fyntrac Model Diagnostic — Tenant: {TENANT}")
print(f"{'='*60}\n")

client = pymongo.MongoClient(MONGO_URI)
db = client[TENANT]

# ── Step 1: Load active model ─────────────────────────────────────────────────
print("[ 1 ] Fetching active model from Models collection...")
model_doc = db["Models"].find_one(
    {"isDeleted": 0, "modelStatus": "ACTIVE"},
    sort=[("orderId", 1)]
)
if not model_doc:
    print("  ✗  No active model found (modelStatus=ACTIVE, isDeleted=0). Aborting.")
    sys.exit(1)

model_name    = model_doc.get("modelName", "<unnamed>")
model_type    = model_doc.get("modelType", "PYTHON")
model_file_id = model_doc.get("modelFileId")
print(f"  ✓  Model: '{model_name}' | type={model_type} | modelFileId={model_file_id}")

# ── Step 2: Load ModelFile ────────────────────────────────────────────────────
print(f"\n[ 2 ] Fetching ModelFile id={model_file_id}...")
if not model_file_id:
    print("  ✗  modelFileId is empty. Aborting.")
    sys.exit(1)

file_doc = db["ModelFiles"].find_one({"_id": ObjectId(model_file_id)})
if not file_doc:
    print(f"  ✗  ModelFile {model_file_id} not found. Aborting.")
    sys.exit(1)

file_data = file_doc.get("fileData")
raw_bytes = bytes(file_data) if not isinstance(file_data, bytes) else file_data
content_type = file_doc.get("contentType", "")
print(f"  ✓  ModelFile loaded: {len(raw_bytes)} bytes | contentType={content_type}")

# ── Step 3: Decode model file → python_code ───────────────────────────────────
print(f"\n[ 3 ] Decoding model file (type={model_type})...")
python_code = ""

if model_type.upper() == "PYTHON":
    python_code = raw_bytes.decode("utf-8")
    print(f"  ✓  Decoded Python source: {len(python_code)} chars")
    print("  --- First 300 chars ---")
    print(python_code[:300])
    print("  ---")
else:
    # EXCEL — read from 'dsl_code' sheet
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as e:
        print(f"  ✗  Failed to open Excel workbook: {e}")
        sys.exit(1)

    print(f"  ✓  Workbook opened. Sheets: {wb.sheetnames}")
    DSL_SHEET = "dsl_code"
    if DSL_SHEET not in wb.sheetnames:
        print(f"  ✗  Sheet '{DSL_SHEET}' not found. Available: {wb.sheetnames}")
        print("  → Trying first sheet as fallback...")
        DSL_SHEET = wb.sheetnames[0]

    ws = wb[DSL_SHEET]
    lines = []
    for row in ws.iter_rows(values_only=True):
        cell_val = row[0] if row else None
        lines.append(str(cell_val) if cell_val is not None else "")
    python_code = "\n".join(lines)
    print(f"  ✓  Extracted {len(lines)} lines from sheet '{DSL_SHEET}'")
    print("  --- First 300 chars ---")
    print(python_code[:300])
    print("  ---")

if not python_code.strip():
    print("  ✗  Extracted Python code is empty. Aborting.")
    sys.exit(1)

# ── Step 4: Compile template ──────────────────────────────────────────────────
print(f"\n[ 4 ] Compiling model template via ModelRunner...")
from app.python_model.model_runner import ModelRunner
runner = ModelRunner()
try:
    exec_globals = runner.compile_template(python_code)
    print(f"  ✓  Compiled OK. exec_globals keys: {[k for k in exec_globals if not k.startswith('__')][:10]}")
except Exception as e:
    print(f"  ✗  Compilation failed: {e}")
    sys.exit(1)

# ── Step 5: Fetch EventHistory ────────────────────────────────────────────────
print(f"\n[ 5 ] Fetching EventHistory from MongoDB (tenant={TENANT})...")
query = {}
if POSTING_DATE:
    query["postingDate"] = POSTING_DATE

all_events = list(db["EventHistory"].find(query).sort("priority", -1))
print(f"  ✓  Found {len(all_events)} events")

if not all_events:
    print("  ✗  No events to process. Nothing to run.")
    sys.exit(0)

# Serialize ObjectIds
for e in all_events:
    if "_id" in e:
        e["_id"] = str(e["_id"])

# Print sample event
sample = all_events[0]
print(f"  Sample event (instrumentId={sample.get('instrumentId')}, postingDate={sample.get('postingDate')}):")
print(f"    keys: {list(sample.keys())}")

# ── Step 6: Transform ─────────────────────────────────────────────────────────
print(f"\n[ 6 ] Running data_transformer...")
from app.python_model.data_transformer import transform

first_event  = all_events[0]
posting_date = first_event.get("postingDate", 20220101)
date_str     = str(posting_date)
posting_date_str = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"

# Group by instrument for testing - just use the first instrument's events
first_instr = first_event.get("instrumentId")
instr_events = [e for e in all_events if e.get("instrumentId") == first_instr]
print(f"  Testing with instrument: {first_instr} ({len(instr_events)} events) postingDate={posting_date_str}")

try:
    event_data_list, raw_event_data = transform(instr_events, posting_date_str)
    print(f"  ✓  Transform OK: {len(event_data_list)} rows")
    if event_data_list:
        print(f"  Sample transformed row keys: {list(event_data_list[0].keys())[:10]}")
except Exception as e:
    print(f"  ✗  Transform failed: {e}")
    import traceback; traceback.print_exc()
    sys.exit(1)

if not event_data_list:
    print("  ✗  Transform produced 0 rows. Check transformer logic.")
    sys.exit(0)

# ── Step 7: Run model ─────────────────────────────────────────────────────────
print(f"\n[ 7 ] Running model for instrument {first_instr}...")
try:
    result = runner.run(
        python_code=python_code,
        event_data=event_data_list,
        raw_event_data=raw_event_data,
        override_postingdate=posting_date_str,
        exec_globals=exec_globals,
    )
except Exception as e:
    print(f"  ✗  ModelRunner.run() raised exception: {e}")
    import traceback; traceback.print_exc()
    sys.exit(1)

print(f"\n{'='*60}")
print(f"  RESULT")
print(f"{'='*60}")

if result.get("error"):
    print(f"  ✗  Model execution error: {result['error']}")
    if "traceback" in result:
        print(result["traceback"])
else:
    transactions = result.get("transactions", [])
    print(f"  ✓  Generated {len(transactions)} transactions")
    for i, tx in enumerate(transactions):
        print(f"\n  Transaction [{i+1}]:")
        for k, v in tx.items():
            print(f"    {k}: {v}")

print(f"\n{'='*60}\n")
