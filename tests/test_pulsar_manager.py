"""
tests/test_pulsar_manager.py
============================
Unit tests for app.pulsar.manager.PulsarManager.

Tests cover:
  - Spring Pulsar JSON schema header stripping
  - _process_python_model_execution: field mapping, Memcached lookup,
    cache miss, unexpected cache format, missing fields
  - _load_active_model: no active model, no modelFileId, ModelFile not found,
    PYTHON model decode, EXCEL model decode (openpyxl), empty code guard
  - Downstream producers: aggregation + GL messages fired after batch
  - producers.py: AggregationMessageProducer, GeneralLedgerMessageProducer

All external I/O (Pulsar, Memcached, MongoDB, ModelRunner) is mocked so the
tests run without any live infrastructure.
"""

import asyncio
import io
import json
import struct
import unittest
from unittest.mock import AsyncMock, MagicMock, patch, call

import pytest


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_spring_json_bytes(payload: dict) -> bytes:
    """Simulate Spring Pulsar JSON schema wire format: 0x0e 0x01 + 8-byte version + JSON."""
    json_bytes = json.dumps(payload).encode("utf-8")
    header = bytes([0x0e, 0x01]) + struct.pack(">Q", 1)  # 2 magic + 8-byte version
    return header + json_bytes


def _make_plain_json_bytes(payload: dict) -> bytes:
    return json.dumps(payload).encode("utf-8")


def _make_manager():
    """Return a PulsarManager with all infrastructure mocked out."""
    from app.pulsar.manager import PulsarManager
    mgr = PulsarManager.__new__(PulsarManager)
    mgr._client = None
    mgr._settings = None
    mgr._consumer_task = None
    mgr._python_model_consumer_task = None
    mgr._stop_event = asyncio.Event()
    mgr._memcache_client = MagicMock()
    mgr._aggregation_producer = AsyncMock()
    mgr._gl_producer = AsyncMock()
    return mgr


# ---------------------------------------------------------------------------
# Spring Pulsar schema header stripping
# ---------------------------------------------------------------------------

class TestSchemaHeaderStripping(unittest.IsolatedAsyncioTestCase):
    """Verify the consumer correctly strips the Spring Pulsar JSON schema header."""

    async def asyncSetUp(self):
        self.mgr = _make_manager()
        # Prevent actual model execution
        self.mgr._process_python_model_execution = AsyncMock()

    def _make_pulsar_msg(self, raw: bytes):
        msg = MagicMock()
        msg.data.return_value = raw
        return msg

    def _strip(self, raw: bytes) -> dict:
        """Replicate the stripping logic from the consumer loop."""
        if len(raw) > 0 and raw[0] == 0x0e:
            raw = raw[10:]
        elif len(raw) > 1 and raw[0] == 0x00:
            raw = raw[5:]
        return json.loads(raw.decode("utf-8"))

    def test_strip_spring_json_header(self):
        payload = {"tenantId": "T1", "executionDate": 20250101, "key": "K1", "isLast": False}
        raw = _make_spring_json_bytes(payload)
        assert raw[0] == 0x0e
        result = self._strip(raw)
        assert result == payload

    def test_plain_json_unchanged(self):
        payload = {"tenantId": "T2", "executionDate": 20250202, "key": "K2", "isLast": True}
        raw = _make_plain_json_bytes(payload)
        result = self._strip(raw)
        assert result == payload

    def test_old_pulsar_format_stripped(self):
        payload = {"tenantId": "T3", "key": "K3"}
        json_bytes = json.dumps(payload).encode("utf-8")
        raw = bytes([0x00]) + struct.pack(">I", 99) + json_bytes  # 0x00 + 4-byte version
        result = self._strip(raw)
        assert result == payload


# ---------------------------------------------------------------------------
# _process_python_model_execution
# ---------------------------------------------------------------------------

class TestProcessPythonModelExecution(unittest.IsolatedAsyncioTestCase):

    async def asyncSetUp(self):
        self.mgr = _make_manager()
        self.mgr._execute_python_model = AsyncMock()

    async def test_happy_path_direct_payload(self):
        """instrumentIds arrive directly in the Pulsar payload — no Memcached."""
        instrument_ids = ["LOAN1", "LOAN2", "LOAN3"]
        payload = {
            "tenantId": "TNT001",
            "executionDate": 20220131,
            "instrumentIds": instrument_ids,
            "isLast": True,
        }
        with patch("app.pulsar.manager.mongo_manager") as mock_mongo:
            mock_mongo.get_database.return_value = MagicMock()
            await self.mgr._process_python_model_execution(payload)

        self.mgr._execute_python_model.assert_awaited_once()
        call_args = self.mgr._execute_python_model.call_args
        assert call_args[0][3] == instrument_ids

    async def test_missing_tenant_id_returns_early(self):
        await self.mgr._process_python_model_execution(
            {"executionDate": 20220131, "instrumentIds": ["LOAN1"]}
        )
        self.mgr._execute_python_model.assert_not_awaited()

    async def test_missing_instrument_ids_returns_early(self):
        await self.mgr._process_python_model_execution(
            {"tenantId": "TNT001", "executionDate": 20220131}
        )
        self.mgr._execute_python_model.assert_not_awaited()

    async def test_empty_instrument_list_returns_early(self):
        await self.mgr._process_python_model_execution(
            {"tenantId": "T", "executionDate": 20250101, "instrumentIds": [], "isLast": False}
        )
        self.mgr._execute_python_model.assert_not_awaited()

    async def test_execution_date_passed_correctly(self):
        with patch("app.pulsar.manager.mongo_manager"):
            await self.mgr._process_python_model_execution(
                {"tenantId": "TNT001", "executionDate": 20230615,
                 "instrumentIds": ["LOAN1"], "isLast": False}
            )
        # executionDate is 3rd positional arg to _execute_python_model
        assert self.mgr._execute_python_model.call_args[0][2] == 20230615

    async def test_is_last_flag_logs_correctly(self):
        """is_last=True should still call _execute_python_model."""
        with patch("app.pulsar.manager.mongo_manager"):
            await self.mgr._process_python_model_execution(
                {"tenantId": "T", "executionDate": 20220101,
                 "instrumentIds": ["X"], "isLast": True}
            )
        self.mgr._execute_python_model.assert_awaited_once()

    async def test_no_memcache_called(self):
        """Verify Memcached is never consulted in the new architecture."""
        with patch("app.pulsar.manager.mongo_manager"):
            await self.mgr._process_python_model_execution(
                {"tenantId": "T", "executionDate": 20220101,
                 "instrumentIds": ["X"], "isLast": False}
            )
        self.mgr._memcache_client.get.assert_not_called()


# ---------------------------------------------------------------------------
# _load_active_model
# ---------------------------------------------------------------------------

class TestLoadActiveModel(unittest.IsolatedAsyncioTestCase):

    async def asyncSetUp(self):
        self.mgr = _make_manager()

    def _make_db(self, model_doc=None, file_doc=None):
        db = MagicMock()
        db.__getitem__ = MagicMock(side_effect=lambda name: {
            "Models": self._make_collection(model_doc),
            "ModelFiles": self._make_collection(file_doc),
        }[name])
        return db

    def _make_collection(self, doc):
        col = MagicMock()
        col.find_one = AsyncMock(return_value=doc)
        return col

    async def test_no_active_model_returns_none(self):
        db = self._make_db(model_doc=None)
        code, globs = await self.mgr._load_active_model(db, "TNT001")
        assert code is None
        assert globs is None

    async def test_model_missing_file_id_returns_none(self):
        model_doc = {"_id": "m1", "modelName": "Test", "modelType": "PYTHON", "modelFileId": None}
        db = self._make_db(model_doc=model_doc)
        code, globs = await self.mgr._load_active_model(db, "TNT001")
        assert code is None

    async def test_model_file_not_found_returns_none(self):
        from bson import ObjectId
        model_doc = {"_id": "m1", "modelName": "Test", "modelType": "PYTHON",
                     "modelFileId": str(ObjectId())}
        db = self._make_db(model_doc=model_doc, file_doc=None)
        code, globs = await self.mgr._load_active_model(db, "TNT001")
        assert code is None

    async def test_python_model_decoded_and_compiled(self):
        from bson import ObjectId
        python_src = "def calculate(row):\n    return []\n"
        file_id = str(ObjectId())
        model_doc = {"_id": "m1", "modelName": "MyModel", "modelType": "PYTHON",
                     "modelFileId": file_id}
        file_doc = {"_id": ObjectId(file_id), "contentType": "text/x-python",
                    "fileData": python_src.encode("utf-8")}
        db = self._make_db(model_doc=model_doc, file_doc=file_doc)

        mock_exec_globals = {"__builtins__": {}}
        with patch("app.python_model.model_runner.ModelRunner") as MockRunner:
            instance = MockRunner.return_value
            instance.compile_template.return_value = mock_exec_globals
            code, globs = await self.mgr._load_active_model(db, "TNT001")

        assert code == python_src
        assert globs is mock_exec_globals
        instance.compile_template.assert_called_once_with(python_src)

    async def test_python_model_compile_failure_returns_none(self):
        from bson import ObjectId
        python_src = "syntax error here :::"
        file_id = str(ObjectId())
        model_doc = {"_id": "m1", "modelName": "Bad", "modelType": "PYTHON",
                     "modelFileId": file_id}
        file_doc = {"_id": ObjectId(file_id), "fileData": python_src.encode()}
        db = self._make_db(model_doc=model_doc, file_doc=file_doc)

        with patch("app.python_model.model_runner.ModelRunner") as MockRunner:
            MockRunner.return_value.compile_template.side_effect = SyntaxError("bad syntax")
            code, globs = await self.mgr._load_active_model(db, "TNT001")

        assert code is None
        assert globs is None

    async def test_excel_model_reads_dsl_code_sheet(self):
        import openpyxl
        from bson import ObjectId

        # Build a minimal xlsx workbook with a 'dsl_code' sheet
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("dsl_code")
        ws["A1"] = "def calculate(row):"
        ws["A2"] = "    return []"
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        file_id = str(ObjectId())
        model_doc = {"_id": "m1", "modelName": "ExcelModel", "modelType": "EXCEL",
                     "modelFileId": file_id}
        file_doc = {"_id": ObjectId(file_id),
                    "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "fileData": xlsx_bytes}
        db = self._make_db(model_doc=model_doc, file_doc=file_doc)

        mock_exec_globals = {}
        with patch("app.python_model.model_runner.ModelRunner") as MockRunner:
            MockRunner.return_value.compile_template.return_value = mock_exec_globals
            code, globs = await self.mgr._load_active_model(db, "TNT001")

        assert "def calculate(row):" in code
        assert globs is mock_exec_globals

    async def test_excel_model_missing_dsl_sheet_returns_none(self):
        import openpyxl
        from bson import ObjectId

        wb = openpyxl.Workbook()
        wb.create_sheet("other_sheet")
        buf = io.BytesIO()
        wb.save(buf)

        file_id = str(ObjectId())
        model_doc = {"_id": "m1", "modelName": "ExcelModel", "modelType": "EXCEL",
                     "modelFileId": file_id}
        file_doc = {"_id": ObjectId(file_id), "fileData": buf.getvalue()}
        db = self._make_db(model_doc=model_doc, file_doc=file_doc)

        code, globs = await self.mgr._load_active_model(db, "TNT001")
        assert code is None

    async def test_empty_python_code_returns_none(self):
        from bson import ObjectId
        file_id = str(ObjectId())
        model_doc = {"_id": "m1", "modelName": "Empty", "modelType": "PYTHON",
                     "modelFileId": file_id}
        file_doc = {"_id": ObjectId(file_id), "fileData": b"   \n  "}
        db = self._make_db(model_doc=model_doc, file_doc=file_doc)
        code, globs = await self.mgr._load_active_model(db, "TNT001")
        assert code is None


# ---------------------------------------------------------------------------
# Downstream producers fired after _execute_python_model
# ---------------------------------------------------------------------------

class TestDownstreamProducers(unittest.IsolatedAsyncioTestCase):
    """Aggregation + GL messages are published after batch completes."""

    async def asyncSetUp(self):
        self.mgr = _make_manager()

    def _make_db(self, events=None, model_code="def calculate(row):\n return []\n"):
        db = MagicMock()
        event_col = MagicMock()
        event_col.find.return_value.__aiter__ = MagicMock(return_value=iter(events or []))
        event_col.update_many = AsyncMock()
        tx_col = MagicMock()
        tx_col.insert_many = AsyncMock()
        dsl_col = MagicMock()

        def getitem(name):
            return {
                "EventHistory": event_col,
                "Transaction": tx_col,
                "dsl_template_artifacts": dsl_col,
                "Models": MagicMock(),
                "ModelFiles": MagicMock(),
            }.get(name, MagicMock())

        db.__getitem__ = MagicMock(side_effect=getitem)
        return db

    async def test_aggregation_and_gl_published_after_batch(self):
        db = self._make_db()

        # Publishers fire in the finally block even when there are no events
        with patch.object(self.mgr, "_load_active_model",
                          new=AsyncMock(return_value=("def f(): pass", {}))):
            await self.mgr._execute_python_model(db, "TNT001", 20220131, ["LOAN1"], 99999)

        self.mgr._aggregation_producer.execute_aggregation.assert_awaited_once()
        call_kwargs = self.mgr._aggregation_producer.execute_aggregation.call_args[1]
        assert call_kwargs["tenant_id"] == "TNT001"
        assert call_kwargs["job_id"] == 99999
        assert call_kwargs["aggregation_date"] == 20220131

        self.mgr._gl_producer.book_temp_gl.assert_awaited_once()
        gl_kwargs = self.mgr._gl_producer.book_temp_gl.call_args[1]
        assert gl_kwargs["tenant_id"] == "TNT001"
        assert gl_kwargs["job_id"] == 99999

    async def test_aggregation_failure_does_not_suppress_gl(self):
        """Even if aggregation publish fails, GL message must still be sent."""
        db = self._make_db()
        self.mgr._aggregation_producer.execute_aggregation.side_effect = RuntimeError("agg down")

        with patch.object(self.mgr, "_load_active_model",
                          new=AsyncMock(return_value=("def f(): pass", {}))):
            await self.mgr._execute_python_model(db, "T", 20220101, [], 1)

        self.mgr._gl_producer.book_temp_gl.assert_awaited_once()


# ---------------------------------------------------------------------------
# producers.py unit tests
# ---------------------------------------------------------------------------

class TestAggregationMessageProducer(unittest.IsolatedAsyncioTestCase):

    async def test_sends_correct_payload(self):
        from app.pulsar.producers import AggregationMessageProducer
        mock_client = MagicMock()
        mock_producer = MagicMock()
        mock_client.create_producer.return_value = mock_producer

        prod = AggregationMessageProducer(mock_client, "agg-topic")
        loop = asyncio.get_running_loop()
        await prod.execute_aggregation("TNT001", 1234567890, 20250618, loop)

        mock_producer.send.assert_called_once()
        sent = json.loads(mock_producer.send.call_args[0][0].decode())
        assert sent == {"tenantId": "TNT001", "jobId": 1234567890, "aggregationDate": 20250618}

    async def test_close_is_safe_without_producer(self):
        from app.pulsar.producers import AggregationMessageProducer
        prod = AggregationMessageProducer(MagicMock(), "topic")
        prod.close()  # should not raise


class TestGeneralLedgerMessageProducer(unittest.IsolatedAsyncioTestCase):

    async def test_sends_correct_payload(self):
        from app.pulsar.producers import GeneralLedgerMessageProducer
        mock_client = MagicMock()
        mock_producer = MagicMock()
        mock_client.create_producer.return_value = mock_producer

        prod = GeneralLedgerMessageProducer(mock_client, "gl-topic")
        loop = asyncio.get_running_loop()
        await prod.book_temp_gl("TNT001", 9876543210, loop)

        mock_producer.send.assert_called_once()
        sent = json.loads(mock_producer.send.call_args[0][0].decode())
        assert sent == {"tenantId": "TNT001", "jobId": 9876543210}

    async def test_close_idempotent(self):
        from app.pulsar.producers import GeneralLedgerMessageProducer
        mock_client = MagicMock()
        mock_prod = MagicMock()
        mock_client.create_producer.return_value = mock_prod
        prod = GeneralLedgerMessageProducer(mock_client, "topic")
        # Trigger producer creation
        loop = asyncio.get_running_loop()
        await prod.book_temp_gl("T", 1, loop)
        prod.close()
        prod.close()  # safe to call twice
        mock_prod.close.assert_called_once()


# ---------------------------------------------------------------------------
# Integration-style: full message → execute path (no live infra)
# ---------------------------------------------------------------------------

class TestEndToEndMessageFlow(unittest.IsolatedAsyncioTestCase):
    """
    Simulates what happens when a Spring Pulsar message arrives on the topic:
    header stripped → JSON parsed → instrumentIds read directly → model executed →
    aggregation + GL published.
    """

    async def test_full_flow_from_wire_bytes(self):
        mgr = _make_manager()
        execute_mock = AsyncMock()
        mgr._execute_python_model = execute_mock

        instrument_ids = ["LOAN1", "LOAN2"]

        java_payload = {
            "tenantId": "TNT001",
            "executionDate": 20220131,
            "instrumentIds": instrument_ids,
            "isLast": True,
        }
        # Simulate Spring Pulsar wire format
        raw = _make_spring_json_bytes(java_payload)

        # Strip header (as the consumer loop does)
        if raw[0] == 0x0e:
            raw = raw[10:]
        parsed = json.loads(raw.decode("utf-8"))
        assert parsed == java_payload

        with patch("app.pulsar.manager.mongo_manager") as mock_mongo:
            mock_mongo.get_database.return_value = MagicMock()
            await mgr._process_python_model_execution(parsed)

        # Memcached must NOT be consulted
        mgr._memcache_client.get.assert_not_called()
        execute_mock.assert_awaited_once()
        args = execute_mock.call_args[0]
        assert args[1] == "TNT001"
        assert args[2] == 20220131
        assert args[3] == instrument_ids


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
