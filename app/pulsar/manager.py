import logging
import json
import asyncio
import os
import threading
import time
from datetime import datetime, timezone
from typing import Optional

import pulsar
from pulsar import Timeout
from pymemcache.client.base import Client as MemcacheClient
from bson.decimal128 import Decimal128

from app.config import Settings
from app.db.mongodb import mongo_manager
from app.pulsar.producers import AggregationMessageProducer, GeneralLedgerMessageProducer

logger = logging.getLogger(__name__)


class PulsarManager:
    """Manages the Pulsar client and background consumer task."""

    def __init__(self):
        self._client: Optional[pulsar.Client] = None
        self._settings: Optional[Settings] = None
        self._consumer_task: Optional[asyncio.Task] = None
        self._python_model_consumer_task: Optional[asyncio.Task] = None
        self._stop_event = asyncio.Event()
        self._memcache_client: Optional[MemcacheClient] = None
        self._aggregation_producer: Optional[AggregationMessageProducer] = None
        self._gl_producer: Optional[GeneralLedgerMessageProducer] = None

    def start(self, settings: Settings) -> None:
        """Initialise Pulsar client and start consumer loop."""
        if not settings.PULSAR_SERVICE_URL:
            logger.warning("PULSAR_SERVICE_URL not set; skipping Pulsar integration.")
            return

        self._settings = settings
        logger.info("Connecting to Pulsar at %s", settings.PULSAR_SERVICE_URL)
        
        try:
            self._client = pulsar.Client(
                settings.PULSAR_SERVICE_URL,
                logger=pulsar.ConsoleLogger(pulsar.LoggerLevel.Error),
            )
            self._stop_event.clear()
            
            # Setup memcached client
            self._memcache_client = MemcacheClient(
                (settings.MEMCACHED_HOST, settings.MEMCACHED_PORT),
                connect_timeout=5,
                timeout=5
            )

            # Instantiate downstream producers (mirror Java AggregationMessageProducer
            # and GeneralLedgerMessageProducer)
            self._aggregation_producer = AggregationMessageProducer(
                self._client, settings.PULSAR_AGGREGATION_TOPIC
            )
            self._gl_producer = GeneralLedgerMessageProducer(
                self._client, settings.PULSAR_GL_STAGING_TOPIC
            )

            # Start the background tasks within the current asyncio event loop
            loop = asyncio.get_running_loop()
            self._consumer_task = loop.create_task(self._consumer_loop())
            self._python_model_consumer_task = loop.create_task(self._python_model_consumer_loop())
            logger.info("Pulsar background consumers (EventHistory + PythonModel) and Memcached connected.")
        except Exception as e:
            logger.error("Failed to start Pulsar/Memcache manager: %s", e)

    async def close(self) -> None:
        """Signal background tasks to stop and close client."""
        logger.info("Stopping Pulsar manager...")
        self._stop_event.set()
        
        for task_name, task in [("EventHistory", self._consumer_task),
                                ("PythonModel", self._python_model_consumer_task)]:
            if task:
                try:
                    await asyncio.wait_for(task, timeout=5.0)
                except asyncio.TimeoutError:
                    logger.warning("%s consumer task took too long. Cancelling...", task_name)
                    task.cancel()
                except Exception as e:
                    logger.error("Error during %s consumer shutdown: %s", task_name, e)
                
        if self._client:
            self._client.close()
            logger.info("Pulsar connection closed.")

        # Close downstream producers
        if self._aggregation_producer:
            self._aggregation_producer.close()
        if self._gl_producer:
            self._gl_producer.close()
            
        if self._memcache_client:
            self._memcache_client.close()
            logger.info("Memcache connection closed.")

    async def _consumer_loop(self):
        """Background loop that polls for messages and processes them."""
        logger.info("Subscribing to topic: %s", self._settings.PULSAR_EVENT_HISTORY_QUERY_TOPIC)
        
        try:
            consumer = self._client.subscribe(
                self._settings.PULSAR_EVENT_HISTORY_QUERY_TOPIC,
                subscription_name=self._settings.PULSAR_SUBSCRIPTION_NAME,
                consumer_type=pulsar.ConsumerType.Shared,
            )
            
            producer = self._client.create_producer(self._settings.PULSAR_EVENT_HISTORY_RESULT_TOPIC)
        except Exception as e:
             logger.error("Failed to create Pulsar consumer/producer: %s", e)
             return

        loop = asyncio.get_running_loop()

        while not self._stop_event.is_set():
            try:
                # receive(timeout_millis) runs in executor to avoid blocking the async event loop
                msg = await loop.run_in_executor(None, consumer.receive, 1000)
            except Timeout:
                # Normal timeout when no messages are available, just loop again
                continue
            except Exception as e:
                logger.error("Error receiving from Pulsar: %s", e)
                await asyncio.sleep(1)
                continue

            try:
                payload = json.loads(msg.data().decode('utf-8'))
                logger.info("Received Pulsar message: tenantId=%s", payload.get("tenantId"))
                
                # Process message async and await the result
                await self._process_query(payload, producer)
                
                # Acknowledge successfully processed message
                consumer.acknowledge(msg)
            except json.JSONDecodeError as e:
                logger.error("Failed to decode message JSON: %s", e)
                consumer.acknowledge(msg) # Ack invalid messages so they are discarded
            except Exception as e:
                logger.error("Error processing Pulsar message: %s", e, exc_info=True)
                # Negative acknowledge to retry
                consumer.negative_acknowledge(msg)
                
        consumer.close()
        producer.close()

    async def _process_query(self, payload: dict, producer: pulsar.Producer):
        """Handle an incoming batched query and publish back results."""
        tenant_id = payload.get("tenantId")
        instrument_ids = payload.get("instrumentIds", [])
        job_id = payload.get("jobId")
        posting_date = payload.get("postingDate")

        if not tenant_id or not instrument_ids:
            logger.warning("Invalid payload: missing tenantId or instrumentIds. Payload: %s", payload)
            return

        # Fetch tenant database
        try:
            db = mongo_manager.get_database(tenant_id)
        except Exception as e:
            logger.error("Failed to get DB for tenant %s: %s", tenant_id, e)
            return

        collection = db["EventHistory"]
        tracker_collection = db["EventHistoryBatchTracker"]
        
        # Mark as PROCESSING
        try:
            from datetime import datetime, timezone
            await tracker_collection.update_one(
                {"jobId": job_id, "tenantId": tenant_id},
                {"$set": {"status": "PROCESSING", "updatedAt": datetime.now(timezone.utc)}}
            )
        except Exception as e:
            logger.warning("Failed to update tracking to PROCESSING for job %s: %s", job_id, e)

        # Build dynamic query
        # Use exact match for $in since there are batched IDs
        query = {"instrumentId": {"$in": instrument_ids}}
        if posting_date is not None:
            query["postingDate"] = posting_date

        logger.info("Querying EventHistory for %d instruments in tenant %s", len(instrument_ids), tenant_id)
        
        # We sort by descending priority just like the REST endpoint
        results = []
        try:
            async for doc in collection.find(query).sort("priority", -1):
                if "_id" in doc:
                    doc["_id"] = str(doc["_id"])  # serialize ObjectId
                results.append(doc)
        except Exception as e:
            logger.error("Database query failed: %s", e)
            try:
                from datetime import datetime, timezone
                await tracker_collection.update_one(
                    {"jobId": job_id, "tenantId": tenant_id},
                    {"$set": {
                        "status": "FAILED", 
                        "errorMessage": str(e),
                        "updatedAt": datetime.now(timezone.utc)
                    }}
                )
            except Exception as update_err:
                logger.error("Failed to mark FAILED status: %s", update_err)
            raise e # Reraise to negatively acknowledge message

        logger.info("Found %d events for batch. Serializing to Memcached.", len(results))

        cache_key = f"event-history-{tenant_id}-{job_id}"
        
        # We push to memcached synchronously or via run_in_executor
        try:
            results_json = json.dumps(results)
            loop = asyncio.get_running_loop()
            # Push to memcached with an expiration of 1 hour (3600 seconds)
            await loop.run_in_executor(None, self._memcache_client.set, cache_key, results_json.encode('utf-8'), 3600)
            logger.info("Successfully pushed results to Memcached at key: %s", cache_key)
        except Exception as e:
            logger.error("Failed to push results to Memcached: %s", e)
            try:
                from datetime import datetime, timezone
                await tracker_collection.update_one(
                    {"jobId": job_id, "tenantId": tenant_id},
                    {"$set": {
                        "status": "FAILED", 
                        "errorMessage": f"Memcached Error: {str(e)}",
                        "updatedAt": datetime.now(timezone.utc)
                    }}
                )
            except Exception as update_err:
                pass
            raise e

        # Mark as COMPLETED in Mongo Tracker
        try:
            from datetime import datetime, timezone
            await tracker_collection.update_one(
                {"jobId": job_id, "tenantId": tenant_id},
                {"$set": {
                    "status": "COMPLETED", 
                    "processedInstrumentCount": len(results),
                    "updatedAt": datetime.now(timezone.utc)
                }}
            )
        except Exception as update_err:
            logger.warning("Failed to mark tracker as COMPLETED for job %s: %s", job_id, update_err)

        # Build result payload
        result_payload = {
            "tenantId": tenant_id,
            "jobId": job_id,
            "success": True,
            "error": None,
            "cacheKey": cache_key,
            "resultCount": len(results)
        }

        # Block to send the result (or run in executor to avoid blocking loop if network is slow)
        try:
            encoded_data = json.dumps(result_payload).encode('utf-8')
            await loop.run_in_executor(None, producer.send, encoded_data)
        except Exception as e:
             logger.error("Failed to publish result to Pulsar: %s", e)
             raise e

    async def _python_model_consumer_loop(self):
        """Background loop that polls for Python model execution messages."""
        topic = self._settings.PULSAR_PYTHON_MODEL_EXECUTION_TOPIC
        subscription = self._settings.PULSAR_PYTHON_MODEL_SUBSCRIPTION_NAME
        logger.info("Subscribing to Python model execution topic: %s", topic)

        try:
            consumer = self._client.subscribe(
                topic,
                subscription_name=subscription,
                consumer_type=pulsar.ConsumerType.Shared,
            )
        except Exception as e:
            logger.error("Failed to create Python model consumer: %s", e)
            return

        loop = asyncio.get_running_loop()

        while not self._stop_event.is_set():
            try:
                msg = await loop.run_in_executor(None, consumer.receive, 1000)
            except Timeout:
                continue
            except Exception as e:
                logger.error("Error receiving Python model message: %s", e)
                await asyncio.sleep(1)
                continue

            try:
                raw = msg.data()
                # Spring Pulsar JSON schema prefixes messages with a 10-byte header:
                # [0x0e, 0x01, <8 bytes schema version>]
                # Strip it so we can parse plain JSON.
                if len(raw) > 0 and raw[0] == 0x0e:
                    raw = raw[10:]
                elif len(raw) > 1 and raw[0] == 0x00:
                    # Older Pulsar schema wire format: 0x00 + 4-byte schema version
                    raw = raw[5:]

                payload = json.loads(raw.decode('utf-8'))
                logger.info(
                    "Received Python model execution message: tenantId=%s, executionDate=%s, key=%s, isLast=%s",
                    payload.get("tenantId"), payload.get("executionDate"),
                    payload.get("key"), payload.get("isLast"),
                )

                await self._process_python_model_execution(payload)

                consumer.acknowledge(msg)
            except json.JSONDecodeError as e:
                logger.error("Failed to decode Python model message JSON: %s | raw_hex=%s", e, msg.data()[:20].hex())
                consumer.acknowledge(msg)  # Ack bad messages so they don’t block the queue
            except Exception as e:
                logger.error("Error processing Python model message: %s", e, exc_info=True)
                consumer.negative_acknowledge(msg)

        consumer.close()

    async def _process_python_model_execution(self, payload: dict):
        """Handle a PythonModelExecutionMessageRecord from Java dataloader.

        Java record fields (Records.PythonModelExecutionMessageRecord):
            tenantId      : String
            executionDate : Integer  (YYYYMMDD)
            instrumentIds : List<String>  — sent directly in the Pulsar payload
            isLast        : boolean
        """
        tenant_id      = payload.get("tenantId")
        execution_date = payload.get("executionDate")
        instrument_ids = payload.get("instrumentIds", [])
        is_last        = payload.get("isLast", False)

        if not tenant_id or not instrument_ids:
            logger.warning(
                "Invalid Python model payload: missing tenantId or instrumentIds. Payload: %s", payload
            )
            return

        logger.info(
            "Python model: tenant=%s executionDate=%s instruments=%d isLast=%s",
            tenant_id, execution_date, len(instrument_ids), is_last,
        )

        # Fetch tenant database
        try:
            db = mongo_manager.get_database(tenant_id)
        except Exception as e:
            logger.error("Failed to get DB for tenant %s: %s", tenant_id, e)
            return

        numeric_job_id = int(time.time() * 1000)
        await self._execute_python_model(db, tenant_id, execution_date, instrument_ids, numeric_job_id)

        if is_last:
            logger.info(
                "Last chunk processed. Tenant=%s, JobId=%s, Date=%s",
                tenant_id, numeric_job_id, execution_date,
            )
            # ── Write EXECUTION_SUMMARY after the final batch ─────────────────────
            # Aggregate all EXECUTION_BATCH documents for this tenant + postingDate
            # to produce a single authoritative summary of the entire Python run.
            await self._write_python_execution_summary(db, tenant_id, execution_date, numeric_job_id)

    async def _execute_python_model(self, db, tenant_id: str, execution_date: int, instrument_ids: list, job_id: int):
        """Execute the Python model logic for a batch of instruments in parallel and log the execution."""
        import time
        from datetime import datetime, timezone
        
        batch_start_time = time.time()
        error_message = None
        success_count = 0
        failed_count = 0
        
        try:
            success_count, failed_count, error_message = await self._execute_python_model_inner(
                db, tenant_id, execution_date, instrument_ids, job_id
            )
        except Exception as e:
            error_message = str(e)
            logger.error("Unhandled exception in _execute_python_model_inner: %s", e, exc_info=True)
        finally:
            duration_ms = int((time.time() - batch_start_time) * 1000)
            status = "FAILED"
            if not error_message:
                status = "SUCCESS" if failed_count == 0 else "PARTIAL_SUCCESS"
                
            log_doc = {
                "jobId": str(job_id),
                "tenantId": tenant_id,
                "postingDate": execution_date,
                "modelType": "PYTHON",
                "logType": "EXECUTION_BATCH",
                "instrumentIds": instrument_ids,
                "instrumentCount": len(instrument_ids),
                "successCount": success_count,
                "failedCount": failed_count,
                "status": status,
                "errorMessage": error_message,
                "durationMs": duration_ms,
                "createdAt": datetime.now(timezone.utc)
            }
            try:
                await db["ModelExecutionBatchLog"].insert_one(log_doc)
                logger.info("Inserted ModelExecutionBatchLog for job %s: status=%s duration=%dms", job_id, status, duration_ms)
            except Exception as log_err:
                logger.error("Failed to insert ModelExecutionBatchLog: %s", log_err)

    async def _write_python_execution_summary(self, db, tenant_id: str, execution_date: int, job_id: int):
        """Aggregate all EXECUTION_BATCH logs for this tenant+postingDate and write one EXECUTION_SUMMARY record.

        Called after the final Pulsar batch (isLast=True) completes. This produces
        the authoritative end-to-end summary of the entire Python model run.
        """
        from datetime import datetime, timezone
        try:
            batch_logs = await db["ModelExecutionBatchLog"].find(
                {
                    "tenantId": tenant_id,
                    "postingDate": execution_date,
                    "logType": "EXECUTION_BATCH",
                    "modelType": "PYTHON",
                }
            ).to_list(length=None)

            total_batches      = len(batch_logs)
            total_instruments  = sum(d.get("instrumentCount", 0) for d in batch_logs)
            total_success      = sum(d.get("successCount",    0) for d in batch_logs)
            total_failed       = sum(d.get("failedCount",     0) for d in batch_logs)
            total_duration_ms  = sum(d.get("durationMs",      0) for d in batch_logs)
            error_messages     = [d["errorMessage"] for d in batch_logs
                                  if d.get("errorMessage") and d["errorMessage"].strip()]

            if total_failed == 0 and not error_messages:
                summary_status = "SUCCESS"
            elif total_success == 0:
                summary_status = "FAILED"
            else:
                summary_status = "PARTIAL_SUCCESS"

            summary_doc = {
                "jobId":           str(job_id),
                "tenantId":        tenant_id,
                "postingDate":     execution_date,
                "modelType":       "PYTHON",
                "logType":         "EXECUTION_SUMMARY",
                "totalBatches":    total_batches,
                "instrumentCount": total_instruments,
                "successCount":    total_success,
                "failedCount":     total_failed,
                "status":          summary_status,
                "errorMessage":    "; ".join(error_messages) if error_messages else None,
                "durationMs":      total_duration_ms,
                "createdAt":       datetime.now(timezone.utc),
            }
            await db["ModelExecutionBatchLog"].insert_one(summary_doc)
            logger.info(
                "Python EXECUTION_SUMMARY written: tenant=%s date=%s batches=%d "
                "instruments=%d success=%d failed=%d status=%s duration=%dms",
                tenant_id, execution_date, total_batches, total_instruments,
                total_success, total_failed, summary_status, total_duration_ms,
            )
        except Exception as e:
            logger.error("Failed to write Python EXECUTION_SUMMARY: %s", e, exc_info=True)


    async def _execute_python_model_inner(self, db, tenant_id: str, execution_date: int, instrument_ids: list, job_id: int):
        """Inner method that performs the actual model logic."""
        collection = db["EventHistory"]
        max_concurrency = min(32, (os.cpu_count() or 1) * 4)
        
        logger.info("Executing Python model for %d instruments in parallel (threads=%d) "
                     "tenant=%s postingDate=%s",
                     len(instrument_ids), max_concurrency, tenant_id, execution_date)

        # 1. Fetch active Models → resolve ModelFile → extract Python code
        #    Mirrors Java: modelDataService.getActiveModels(tenantId)
        #                  modelDataService.getModelFile(model.getModelFileId(), tenantId)
        python_code = ""
        exec_globals = None
        try:
            python_code, exec_globals = await self._load_active_model(db, tenant_id)
            if not python_code or exec_globals is None:
                return (0, len(instrument_ids), "Active model not found or failed to compile.")
        except Exception as e:
            logger.error("Unexpected error loading model for tenant %s: %s", tenant_id, e)
            return (0, len(instrument_ids), f"Unexpected error loading model: {e}")

        # 2. Fetch all events for the entire chunk in a single query
        logger.info("Fetching data for %d instruments from MongoDB...", len(instrument_ids))
        query = {"instrumentId": {"$in": instrument_ids}}
        if execution_date is not None:
            query["postingDate"] = execution_date

        all_events = []
        loop = asyncio.get_running_loop()

        # ── All processing wrapped in try/finally so downstream publishers
        # always fire, mirroring Java's ModelExecutionService.executeExcelModels finally block.
        try:
            try:
                async for doc in collection.find(query).sort("priority", -1):
                    if "_id" in doc:
                        doc["_id"] = str(doc["_id"])
                    all_events.append(doc)
            except Exception as e:
                logger.error("Error fetching data for chunk: %s", e)
                return (0, len(instrument_ids), f"Error fetching data: {e}")

            if not all_events:
                logger.info("No events found for the given instruments in tenant %s", tenant_id)
                return (0, 0, None)

            # Prepare date string in YYYY-MM-DD format
            date_str = str(execution_date)
            if len(date_str) == 8:
                posting_date_str = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
            else:
                # Fallback for non-8-digit dates (e.g. 2022028)
                try:
                    # Try to parse as YYYYMMDD by padding if needed, 
                    # but be careful not to corrupt the year.
                    if len(date_str) == 7:
                        # 2022028 -> 2022-02-08
                        posting_date_str = f"{date_str[:4]}-{date_str[4:6]}-0{date_str[6:]}"
                    else:
                        posting_date_str = date_str
                except:
                    posting_date_str = date_str

            # 3. Transform the data ONCE for the whole chunk
            from app.python_model.data_transformer import transform
            try:
                event_data_list, raw_event_data = await loop.run_in_executor(
                    None, 
                    transform, 
                    all_events, 
                    posting_date_str
                )
            except Exception as e:
                logger.error("Failed to transform data for chunk: %s", e)
                return (0, len(instrument_ids), f"Failed to transform data: {e}")

            # Map event_data and raw event docs by instrumentid for O(1) lookup
            instrument_data_map = {}
            instrument_event_map = {}   # first EventHistory doc per instrument for metadata
            for row in event_data_list:
                iid = row.get("instrumentid")
                if iid:
                    instrument_data_map[iid] = row
            for doc in all_events:
                iid = doc.get("instrumentId")
                if iid and iid not in instrument_event_map:
                    instrument_event_map[iid] = doc
            
            # 4. Fetch Attributes definitions once for the batch, then resolve
            #    the active InstrumentAttribute values per instrument.
            #    Done here (async, before the thread pool) to keep DB calls off threads.
            attr_definitions = await self._fetch_attribute_definitions(db)
            instrument_attributes_map: dict = {}
            if attr_definitions:
                for _iid in instrument_ids:
                    attr_dict, version_id = await self._fetch_instrument_attributes(
                        db, _iid, attr_definitions
                    )
                    instrument_attributes_map[_iid] = {
                        "attributes": attr_dict,
                        "versionId": version_id
                    }
                    logger.info(
                        "Resolved %d attributes for instrument %s (versionId=%s)",
                        len(attr_dict), _iid, version_id
                    )
            else:
                logger.warning(
                    "No attribute definitions found in Attributes collection for tenant %s; "
                    "TransactionActivity.attributes will be empty.",
                    tenant_id,
                )

            # 5. Run model execution in PARALLEL via ThreadPoolExecutor.
            #    Thread-safety is guaranteed by:
            #      a) dsl_functions.py globals → threading.local() (per-thread state)
            #      b) exec_globals → each thread compiles its OWN template from python_code
            #    Each thread receives python_code (not exec_globals) so model_runner
            #    calls compile_template() per-thread, giving fully isolated namespaces.
            success_count = 0
            import concurrent.futures

            with concurrent.futures.ThreadPoolExecutor(max_workers=max_concurrency) as pool:
                process_futures = []

                for instrument_id in instrument_ids:
                    instr_data = instrument_data_map.get(instrument_id)
                    if not instr_data:
                        # No data for this instrument
                        continue

                    # Enrich instr_data with active InstrumentAttribute values
                    ia_dict = instrument_attributes_map.get(instrument_id, {}).get("attributes", {})
                    for k, v in ia_dict.items():
                        attr_key = f"ATTRIBUTE_{k}"
                        if attr_key not in instr_data:
                            instr_data[attr_key] = v
                        if k not in instr_data:
                            instr_data[k] = v

                    future = loop.run_in_executor(
                        pool,
                        self._process_instrument_pretransformed,
                        tenant_id,
                        instrument_id,
                        posting_date_str,
                        instr_data,
                        raw_event_data,
                        None,           # exec_globals=None → each thread compiles its own
                        python_code,
                        instrument_event_map.get(instrument_id, {}),
                    )
                    process_futures.append(future)

                acct_periods_cache = {}

                if process_futures:
                    results = await asyncio.gather(*process_futures, return_exceptions=True)

                    for r in results:
                        if isinstance(r, Exception):
                            logger.error("Thread pool execution error: %s", r)
                        else:
                            instrument_id, success, transactions, event_doc = r
                            new_status = "COMPLETED" if success else "ERROR_OUT"

                            if success:
                                success_count += 1

                            # Update the EventHistory status in MongoDB
                            try:
                                update_query = {"instrumentId": instrument_id}
                                if execution_date is not None:
                                    update_query["postingDate"] = execution_date

                                await collection.update_many(
                                    update_query,
                                    {"$set": {"status": new_status}}
                                )
                            except Exception as db_err:
                                logger.error("Failed to update status to %s for instrument %s: %s",
                                             new_status, instrument_id, db_err)

                            # Save TransactionActivity documents to MongoDB
                            if transactions:
                                try:
                                    docs = []
                                    for t in transactions:
                                        # Check zero-amount transactions
                                        is_zero = False
                                        try:
                                            if float(t.get("amount", 0)) == 0:
                                                is_zero = True
                                        except (TypeError, ValueError):
                                            pass

                                        ia_data = instrument_attributes_map.get(instrument_id, {})
                                        version_id = ia_data.get("versionId", 0)

                                        doc = self._build_transaction_activity(
                                            t, tenant_id, instrument_id, job_id, version_id
                                        )

                                        # Log and discard zero-amount transactions
                                        if is_zero:
                                            logger.info("Discarding zero-amount TransactionActivity for %s: %s", instrument_id, doc)
                                            continue

                                        # Enrich with context
                                        period_id = doc.get("originalPeriodId", 0)
                                        if period_id not in acct_periods_cache:
                                            try:
                                                acct_periods_cache[period_id] = await db["AccountingPeriod"].find_one({"periodId": period_id})
                                            except Exception as e:
                                                logger.error("Failed to fetch accounting period for periodId %s: %s", period_id, e)
                                                acct_periods_cache[period_id] = None

                                        doc["accountingPeriod"] = acct_periods_cache.get(period_id)
                                        if event_doc:
                                            doc["sourceId"]         = str(event_doc.get("_id", ""))
                                        # Populate attributes from Attributes + InstrumentAttribute
                                        doc["attributes"] = ia_data.get("attributes", {})
                                        docs.append(doc)

                                    if docs:
                                        for d in docs:
                                            logger.info("Generated TransactionActivity doc: %s", d)

                                        await db["TransactionActivity"].insert_many(docs)
                                        logger.info(
                                            "Saved %d TransactionActivity docs for instrument %s",
                                            len(docs), instrument_id,
                                        )
                                    else:
                                        logger.info(
                                            "No non-zero TransactionActivity docs to save for instrument %s",
                                            instrument_id,
                                        )
                                except Exception as tx_err:
                                    logger.error(
                                        "Failed to save TransactionActivity for instrument %s: %s",
                                        instrument_id, tx_err,
                                    )

            logger.info(
                "Python model: completed batch. Instruments=%d, Succeeded=%d, Failed=%d, "
                "Tenant=%s, PostingDate=%s JobId=%s",
                len(instrument_ids), success_count, len(process_futures) - success_count,
                tenant_id, execution_date, job_id
            )
            
            final_success_count = success_count
            final_failed_count = len(process_futures) - success_count

        finally:
            # ── Mirror Java ModelExecutionService finally block ───────────────
            # Publish aggregation trigger → consumed by Java AggregationService
            # Publish GL booking trigger  → consumed by Java GL service
            # Both happen regardless of success, partial failure, or empty event set.
            try:
                await self._aggregation_producer.execute_aggregation(
                    tenant_id=tenant_id,
                    job_id=job_id,
                    aggregation_date=execution_date,  # YYYYMMDD int, same as postingDate
                    loop=loop,
                )
            except Exception as agg_err:
                logger.error("Failed to publish aggregation message for jobId=%s: %s", job_id, agg_err)

            try:
                await self._gl_producer.book_temp_gl(
                    tenant_id=tenant_id,
                    job_id=job_id,
                    loop=loop,
                )
            except Exception as gl_err:
                logger.error("Failed to publish GL staging message for jobId=%s: %s", job_id, gl_err)

        return (final_success_count, final_failed_count, None)

    def _deserialize_cache_list(self, raw: bytes, cache_key: str):
        """Deserialize a Memcached value written by Java's memcachedRepository.

        Java stores objects using Java object serialization (magic 0xac 0xed).
        CacheList<String> deserializes to an object whose 'list' field holds the IDs.

        Falls back to JSON for plain-text / test payloads.

        Returns:
            list[str] on success, or None on failure (error already logged).
        """
        JAVA_MAGIC = b'\xac\xed'

        if isinstance(raw, str):
            raw = raw.encode('utf-8')

        # ── Strategy 1: Java binary serialization ─────────────────────────
        if raw[:2] == JAVA_MAGIC:
            try:
                import javaobj
                obj = javaobj.loads(raw)
                # CacheList extends ArrayList / has a 'list' field or is itself iterable
                ids = self._extract_ids_from_java_obj(obj)
                if ids is not None:
                    logger.info(
                        "Deserialized Java CacheList from Memcached key %s: %d IDs",
                        cache_key, len(ids),
                    )
                    return ids
                logger.error(
                    "Could not extract ID list from Java object for key %s. "
                    "obj type=%s repr=%s", cache_key, type(obj), repr(obj)[:200]
                )
                return None
            except Exception as e:
                logger.error(
                    "javaobj deserialization failed for key %s: %s", cache_key, e
                )
                return None

        # ── Strategy 2: JSON fallback (plain-text / test payloads) ────────
        try:
            decoded = json.loads(raw.decode('utf-8'))
            if isinstance(decoded, list):
                return decoded
            if isinstance(decoded, dict) and "list" in decoded:
                return decoded["list"]
            logger.error(
                "Unexpected JSON structure in Memcached key %s: %s", cache_key, type(decoded)
            )
            return None
        except Exception as e:
            logger.error(
                "Failed to decode Memcached value for key %s "
                "(not Java serial, not JSON). First 20 bytes: %s — %s",
                cache_key, raw[:20].hex(), e,
            )
            return None

    @staticmethod
    def _extract_ids_from_java_obj(obj) -> list:
        """Recursively walk a javaobj-deserialized object to extract string IDs.

        CacheList<String> is serialized with its ArrayList superclass data,
        so the strings are in obj.annotations or accessible via iteration.
        """
        import javaobj

        # javaobj v0.4+ returns JavaObject; try common patterns
        ids = []

        # Pattern A: object has a 'list' field (Jackson-style)
        if hasattr(obj, 'list') and obj.list is not None:
            for item in obj.list:
                ids.append(str(item))
            return ids

        # Pattern B: JavaObject whose classdesc is ArrayList or CacheList
        # — string children are in obj.annotations (class data annotations)
        if hasattr(obj, 'annotations') and obj.annotations:
            for item in obj.annotations:
                if isinstance(item, str):
                    ids.append(item)
                elif hasattr(item, '__class__') and hasattr(item, 'annotations'):
                    # nested JavaObject (String wrapper)
                    pass
            if ids:
                return ids

        # Pattern C: JavaList (javaobj wraps ArrayList-like as a Python list)
        if isinstance(obj, (list, tuple)):
            return [str(i) for i in obj]

        # Pattern D: iterate via __iter__
        try:
            for item in obj:
                if isinstance(item, str):
                    ids.append(item)
                else:
                    ids.append(str(item))
            if ids:
                return ids
        except TypeError:
            pass

        return None

    async def _load_active_model(self, db, tenant_id: str) -> tuple:
        """Fetch the first active Python model from MongoDB and compile it.

        Two-step lookup mirroring Java ModelExecutionService:
          1. Query ``Models`` collection for active, non-deleted records.
          2. Use ``modelFileId`` to fetch the file from ``ModelFiles``.
          3. Decode the BSON Binary ``fileData``:
               - ModelType.PYTHON  → raw bytes are the Python source code (.py file).
               - ModelType.EXCEL   → open as openpyxl workbook, read the "dsl_code" sheet.

        Returns:
            (python_code: str, exec_globals: dict) on success.
            (None, None) on any failure (errors are logged).
        """
        # ── Step 1: Find active models ────────────────────────────────────
        # Java equivalent:
        #   query.addCriteria(Criteria.where("isDeleted").is(0)
        #                              .and("modelStatus").is(ModelStatus.ACTIVE))
        try:
            model_doc = await db["Models"].find_one(
                {"isDeleted": 0, "modelStatus": "ACTIVE"},
                sort=[("orderId", 1)],   # honour orderId ordering like Java
            )
        except Exception as e:
            logger.error("Failed to query Models collection for tenant %s: %s", tenant_id, e)
            return None, None

        if not model_doc:
            logger.error(
                "No active model found in Models collection for tenant %s. "
                "Ensure at least one model has modelStatus='ACTIVE' and isDeleted=0.",
                tenant_id,
            )
            return None, None

        model_file_id = model_doc.get("modelFileId")
        model_name    = model_doc.get("modelName", "<unnamed>")
        model_type    = model_doc.get("modelType", "PYTHON")   # "PYTHON" | "EXCEL"

        logger.info(
            "Found active model '%s' (id=%s, type=%s, modelFileId=%s) for tenant %s",
            model_name, model_doc.get("_id"), model_type, model_file_id, tenant_id,
        )

        if not model_file_id:
            logger.error("Model '%s' has no modelFileId. Cannot load file.", model_name)
            return None, None

        # ── Step 2: Fetch the ModelFile ───────────────────────────────────
        # Java equivalent: modelDataService.getModelFile(model.getModelFileId(), tenantId)
        try:
            from bson import ObjectId
            file_doc = await db["ModelFiles"].find_one({"_id": ObjectId(model_file_id)})
        except Exception as e:
            logger.error(
                "Failed to fetch ModelFile for modelFileId=%s (model='%s'): %s",
                model_file_id, model_name, e,
            )
            return None, None

        if not file_doc:
            logger.error(
                "ModelFile not found for id=%s (model='%s').", model_file_id, model_name
            )
            return None, None

        # ── Step 3: Decode fileData Binary ────────────────────────────────
        # Motor/PyMongo decodes org.bson.types.Binary → bytes automatically.
        file_data = file_doc.get("fileData")
        if file_data is None:
            logger.error("ModelFile %s has no fileData.", model_file_id)
            return None, None

        # Motor returns BSON Binary as bytes directly; handle both just in case.
        raw_bytes: bytes = bytes(file_data) if not isinstance(file_data, bytes) else file_data

        python_code: str = ""

        if model_type.upper() in ["PYTHON", "DSL"]:
            # The file is a raw .py file uploaded via /api/model/upload/python
            # Java: ExcelFileUtil.convertToMongoBinary(dslTemplate) → dslTemplate.getBytes()
            try:
                python_code = raw_bytes.decode("utf-8")
                logger.info(
                    "Decoded Python source (%d bytes) from ModelFile %s.",
                    len(raw_bytes), model_file_id,
                )
            except UnicodeDecodeError as e:
                logger.error("Failed to UTF-8 decode Python model file %s: %s", model_file_id, e)
                return None, None
        else:
            # EXCEL model — open workbook and read the DSL code sheet
            # Sheet name used by ExcelFileService for the generated Python code.
            DSL_CODE_SHEET = "dsl_code"
            try:
                import io
                import openpyxl
                workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
                if DSL_CODE_SHEET not in workbook.sheetnames:
                    logger.error(
                        "Excel model '%s' has no '%s' sheet. Available sheets: %s",
                        model_name, DSL_CODE_SHEET, workbook.sheetnames,
                    )
                    return None, None
                sheet = workbook[DSL_CODE_SHEET]
                # Code is stored one line per row in column A
                lines = []
                for row in sheet.iter_rows(values_only=True):
                    cell_val = row[0] if row else None
                    lines.append(str(cell_val) if cell_val is not None else "")
                python_code = "\n".join(lines)
                logger.info(
                    "Extracted DSL code (%d lines) from sheet '%s' of model '%s'.",
                    len(lines), DSL_CODE_SHEET, model_name,
                )
            except Exception as e:
                logger.error(
                    "Failed to read Excel model '%s' (ModelFile %s): %s",
                    model_name, model_file_id, e,
                )
                return None, None

        if not python_code.strip():
            logger.error("Extracted Python code is empty for model '%s'.", model_name)
            return None, None

        # ── Step 4: Compile the template ─────────────────────────────────
        try:
            from app.python_model.model_runner import ModelRunner
            runner = ModelRunner()
            exec_globals = runner.compile_template(python_code)
            logger.info("Successfully compiled model '%s' for tenant %s.", model_name, tenant_id)
            return python_code, exec_globals
        except Exception as e:
            logger.error("Failed to compile model '%s': %s", model_name, e)
            return None, None

    def _process_instrument_pretransformed(
        self,
        tenant_id: str,
        instrument_id: str,
        posting_date_str: str,
        instr_data: dict,
        raw_event_data: dict,
        exec_globals: dict,
        python_code: str,
        event_doc: dict = None,
    ) -> tuple:
        """Synchronous method executed in a separate thread. Runs the model on pre-transformed data."""
        try:
            from app.python_model.model_runner import ModelRunner
            runner = ModelRunner()

            result = runner.run(
                python_code=python_code,
                event_data=[instr_data],
                raw_event_data=raw_event_data,
                override_postingdate=posting_date_str,
                exec_globals=exec_globals,
            )

            if result.get("error"):
                logger.error("Model execution error for instrument %s: %s", instrument_id, result["error"])
                return (instrument_id, False, [], event_doc or {}, job_id)

            transactions = result.get("transactions", [])
            logger.info("Model executed for %s: generated %d transactions", instrument_id, len(transactions))

            return (instrument_id, True, transactions, event_doc or {})
        except Exception as e:
            logger.error("Error in thread processing instrument %s: %s", instrument_id, e, exc_info=True)
            return (instrument_id, False, [], event_doc or {})

    # ------------------------------------------------------------------
    # Attribute helpers
    # ------------------------------------------------------------------
    async def _fetch_attribute_definitions(self, db) -> list:
        """Fetch all Attribute definition documents from the Attributes collection.

        Actual Attributes schema (from MongoDB):
            {
              _id           : ObjectId,
              attributeName : str,       # e.g. 'MERCHANT_INDUSTRY'
              isReclassable : 0|1,
              isVersionable : 0|1,
              dataType      : str,
              userField     : str,
              ...
            }
        """
        try:
            definitions = []
            async for doc in db["Attributes"].find({}):
                definitions.append(doc)
            logger.info(
                "Loaded %d attribute definitions from Attributes collection.",
                len(definitions),
            )
            if not definitions:
                logger.warning(
                    "Attributes collection is empty — no attribute definitions to process."
                )
            return definitions
        except Exception as e:
            logger.warning("Failed to fetch Attributes definitions: %s", e)
            return []

    async def _fetch_instrument_attributes(self, db, instrument_id: str, attr_definitions: list) -> tuple[dict, int]:
        """Resolve active attribute values for a single instrument.

        Schema (confirmed from MongoDB)
        --------------------------------
        Attributes collection:
            { attributeName, isReclassable: 0|1, isVersionable: 0|1, ... }

        InstrumentAttribute collection — ONE active record per instrument:
            {
              instrumentId : str,
              attributeId  : '1.0',   # NOT linked to Attributes._id
              endDate      : None,    # null = active
              attributes   : {        # flat { attributeName: value } dict
                'MERCHANT_INDUSTRY': 'SPORTING GOODS AND OUTDOORS',
                'INTEREST_RATE': 8.0,
                ...
              }
            }

        Algorithm
        ---------
        1. Fetch the single active InstrumentAttribute for this instrument
           (filter: instrumentId + endDate=null — no attributeId filter).
        2. From qualifying attr_definitions (isReclassable=1 OR isVersionable=1),
           pick ia_doc['attributes'][attributeName].
        3. Return { attributeName: value } dict.
        """
        attr_dict: dict = {}

        # ── Step 1: Fetch the single active InstrumentAttribute ───────────
        # There is only ONE active record per instrument (endDate=null).
        # attributeId on InstrumentAttribute ('1.0') is NOT a join key to Attributes.
        try:
            ia_doc = await db["InstrumentAttribute"].find_one({
                "instrumentId": instrument_id,
                "endDate":      None,
            })
        except Exception as e:
            logger.warning(
                "Failed to query InstrumentAttribute for instrument=%s: %s",
                instrument_id, e,
            )
            return attr_dict, 0

        if ia_doc is None:
            logger.info(
                "No active InstrumentAttribute found for instrument=%s (endDate=null)",
                instrument_id,
            )
            return attr_dict, 0

        ia_attributes = ia_doc.get("attributes")
        if not isinstance(ia_attributes, dict):
            logger.warning(
                "InstrumentAttribute for instrument=%s has no 'attributes' sub-document "
                "(got type=%s); skipping.",
                instrument_id, type(ia_attributes).__name__,
            )
            return attr_dict, ia_doc.get("versionId", 0)

        # ── Step 2: Pick values for qualifying attribute definitions ──────
        for attr_def in attr_definitions:
            if not bool(attr_def.get("isReclassable", 0)):
                continue

            attr_name: str = attr_def.get("attributeName") or ""
            if not attr_name:
                continue

            if attr_name not in ia_attributes:
                logger.info(
                    "attributeName '%s' not in InstrumentAttribute.attributes for instrument=%s; skipping.",
                    attr_name, instrument_id,
                )
                continue

            attr_dict[attr_name] = ia_attributes[attr_name]
            logger.info(
                "Resolved attribute '%s'=%r for instrument %s",
                attr_name, ia_attributes[attr_name], instrument_id,
            )

        return attr_dict, ia_doc.get("versionId", 0)

    @staticmethod
    def _build_transaction_activity(tx: dict, tenant_id: str, instrument_id: str, job_id: int, version_id: int = 0) -> dict:
        """Map ModelRunner transaction output → TransactionActivity MongoDB document.

        ModelRunner fields:             TransactionActivity fields:
          transactiontype          →    transactionName
          amount                   →    amount  (stored as string "%.4f")
          subinstrumentid          →    attributeId
          postingdate  (YYYY-MM-DD)→    postingDate  (YYYYMMDD int)
          effectivedate(YYYY-MM-DD)→    effectiveDate (YYYYMMDD int)
          instrumentid             →    instrumentId
        """
        def _date_to_int(val: str) -> int:
            """Convert YYYY-MM-DD string to YYYYMMDD int."""
            try:
                return int(str(val).replace("-", "")[:8])
            except Exception:
                return 0

        posting_int  = _date_to_int(tx.get("postingdate",  ""))
        effective_int = _date_to_int(tx.get("effectivedate", ""))
        amount_raw   = tx.get("amount", 0)

        return {
            "instrumentId":               instrument_id,
            "transactionName":            tx.get("transactiontype", "").upper(),
            "amount":                     Decimal128(str(float(amount_raw))),
            "attributeId":                tx.get("subinstrumentid", "1.0"),
            "originalPeriodId":           posting_int // 100,   # YYYYMM
            "instrumentAttributeVersionId": version_id,
            "accountingPeriod":           None,  # enriched below if available
            "periodId":                   0,
            "batchId":                    job_id,
            "source":                     "MODEL",
            "sourceId":                   "",
            "postingDate":                posting_int,
            "effectiveDate":              effective_int,
            "attributes":                 {},
            "isReplayable":               0,
            "_class":                     "com.fyntrac.common.entity.TransactionActivity",
        }



# Create singleton instance
pulsar_manager = PulsarManager()
